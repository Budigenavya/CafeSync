"""
==========================================================
                CafeSync POS
                export.py
                PART 1
==========================================================
"""

import os
from datetime import datetime

from flask import (
    Blueprint,
    jsonify,
    send_file
)

from openpyxl import Workbook

from models import fetch_all

export_bp = Blueprint(
    "export",
    __name__
)

EXPORT_FOLDER = "exports"

os.makedirs(
    EXPORT_FOLDER,
    exist_ok=True
)


# ==========================================================
# EXPORT PRODUCTS
# ==========================================================

@export_bp.route(
    "/export/products",
    methods=["GET"]
)
def export_products():

    rows = fetch_all("""

        SELECT

            id,

            name,

            price,

            stock,

            barcode

        FROM products

        ORDER BY name

    """)

    wb = Workbook()

    ws = wb.active

    ws.title = "Products"

    ws.append([

        "ID",

        "Product",

        "Price",

        "Stock",

        "Barcode"

    ])

    for row in rows:

        ws.append([

            row["id"],

            row["name"],

            row["price"],

            row["stock"],

            row["barcode"]

        ])

    filename = os.path.join(

        EXPORT_FOLDER,

        f"products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    )

    wb.save(filename)

    return send_file(

        filename,

        as_attachment=True

    )


# ==========================================================
# EXPORT CATEGORIES
# ==========================================================

@export_bp.route(
    "/export/categories",
    methods=["GET"]
)
def export_categories():

    rows = fetch_all("""

        SELECT

            id,

            name

        FROM categories

        ORDER BY name

    """)

    wb = Workbook()

    ws = wb.active

    ws.title = "Categories"

    ws.append([

        "ID",

        "Category"

    ])

    for row in rows:

        ws.append([

            row["id"],

            row["name"]

        ])

    filename = os.path.join(

        EXPORT_FOLDER,

        f"categories_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    )

    wb.save(filename)

    return send_file(

        filename,

        as_attachment=True

    )


# ==========================================================
# EXPORT INVENTORY
# ==========================================================

@export_bp.route(
    "/export/inventory",
    methods=["GET"]
)
def export_inventory():

    rows = fetch_all("""

        SELECT

            p.name,

            c.name AS category,

            p.price,

            p.stock

        FROM products p

        LEFT JOIN categories c

        ON p.category_id=c.id

        ORDER BY c.name,p.name

    """)

    wb = Workbook()

    ws = wb.active

    ws.title = "Inventory"

    ws.append([

        "Product",

        "Category",

        "Price",

        "Stock"

    ])

    for row in rows:

        ws.append([

            row["name"],

            row["category"],

            row["price"],

            row["stock"]

        ])

    filename = os.path.join(

        EXPORT_FOLDER,

        f"inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    )

    wb.save(filename)

    return send_file(

        filename,

        as_attachment=True

    )
# ==========================================================
# EXPORT DAILY SALES
# ==========================================================

@export_bp.route(
    "/export/daily-sales",
    methods=["GET"]
)
def export_daily_sales():

    rows = fetch_all("""

        SELECT

            id,

            total,

            gst,

            payment_method,

            payment_status,

            order_status,

            created_at

        FROM orders

        WHERE DATE(created_at)=DATE('now')

        ORDER BY created_at DESC

    """)

    wb = Workbook()

    ws = wb.active

    ws.title = "Daily Sales"

    ws.append([

        "Order ID",

        "Total",

        "GST",

        "Payment Method",

        "Payment Status",

        "Order Status",

        "Date"

    ])

    total_sales = 0

    for row in rows:

        ws.append([

            row["id"],

            row["total"],

            row["gst"],

            row["payment_method"],

            row["payment_status"],

            row["order_status"],

            row["created_at"]

        ])

        total_sales += row["total"]

    ws.append([])

    ws.append([

        "",

        "Total Sales",

        total_sales

    ])

    filename = os.path.join(

        EXPORT_FOLDER,

        f"daily_sales_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    )

    wb.save(filename)

    return send_file(

        filename,

        as_attachment=True

    )


# ==========================================================
# EXPORT MONTHLY SALES
# ==========================================================

@export_bp.route(
    "/export/monthly-sales",
    methods=["GET"]
)
def export_monthly_sales():

    rows = fetch_all("""

        SELECT

            id,

            total,

            gst,

            payment_method,

            payment_status,

            order_status,

            created_at

        FROM orders

        WHERE strftime('%Y-%m',created_at)=

              strftime('%Y-%m','now')

        ORDER BY created_at DESC

    """)

    wb = Workbook()

    ws = wb.active

    ws.title = "Monthly Sales"

    ws.append([

        "Order ID",

        "Total",

        "GST",

        "Payment Method",

        "Payment Status",

        "Order Status",

        "Date"

    ])

    total_sales = 0

    for row in rows:

        ws.append([

            row["id"],

            row["total"],

            row["gst"],

            row["payment_method"],

            row["payment_status"],

            row["order_status"],

            row["created_at"]

        ])

        total_sales += row["total"]

    ws.append([])

    ws.append([

        "",

        "Monthly Total",

        total_sales

    ])

    filename = os.path.join(

        EXPORT_FOLDER,

        f"monthly_sales_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    )

    wb.save(filename)

    return send_file(

        filename,

        as_attachment=True

    )


# ==========================================================
# EXPORT ORDERS
# ==========================================================

@export_bp.route(
    "/export/orders",
    methods=["GET"]
)
def export_orders():

    rows = fetch_all("""

        SELECT

            id,

            customer_name,

            table_number,

            total,

            gst,

            payment_method,

            payment_status,

            order_status,

            created_at

        FROM orders

        ORDER BY created_at DESC

    """)

    wb = Workbook()

    ws = wb.active

    ws.title = "Orders"

    ws.append([

        "Order ID",

        "Customer",

        "Table",

        "Total",

        "GST",

        "Payment",

        "Payment Status",

        "Order Status",

        "Date"

    ])

    for row in rows:

        ws.append([

            row["id"],

            row["customer_name"],

            row["table_number"],

            row["total"],

            row["gst"],

            row["payment_method"],

            row["payment_status"],

            row["order_status"],

            row["created_at"]

        ])

    filename = os.path.join(

        EXPORT_FOLDER,

        f"orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    )

    wb.save(filename)

    return send_file(

        filename,

        as_attachment=True

    )
# ==========================================================
# REPORTLAB IMPORTS
# ==========================================================

from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet


styles = getSampleStyleSheet()


# ==========================================================
# EXPORT INVOICE PDF
# ==========================================================

@export_bp.route(
    "/export/invoice/<int:order_id>",
    methods=["GET"]
)
def export_invoice(order_id):

    order = fetch_one("""

        SELECT *

        FROM orders

        WHERE id=?

    """,(order_id,))

    if not order:

        return jsonify({

            "success":False,

            "message":"Order not found"

        }),404

    items = fetch_all("""

        SELECT

            p.name,

            oi.quantity,

            oi.price

        FROM order_items oi

        JOIN products p

        ON oi.product_id=p.id

        WHERE oi.order_id=?

    """,(order_id,))

    filename = os.path.join(

        EXPORT_FOLDER,

        f"invoice_{order_id}.pdf"

    )

    doc = SimpleDocTemplate(

        filename,

        pagesize=A4

    )

    elements = []

    elements.append(

        Paragraph(

            "<b>CafeSync POS</b>",

            styles["Title"]

        )

    )

    elements.append(

        Paragraph(

            f"Invoice No : {order_id}",

            styles["Normal"]

        )

    )

    elements.append(

        Paragraph(

            f"Date : {order['created_at']}",

            styles["Normal"]

        )

    )

    elements.append(

        Paragraph("<br/>",styles["Normal"])

    )

    data = [

        [

            "Item",

            "Qty",

            "Price",

            "Total"

        ]

    ]

    grand_total = 0

    for item in items:

        total = item["quantity"] * item["price"]

        grand_total += total

        data.append([

            item["name"],

            item["quantity"],

            item["price"],

            total

        ])

    data.append([

        "",

        "",

        "GST",

        order["gst"]

    ])

    data.append([

        "",

        "",

        "Grand Total",

        order["total"]

    ])

    table = Table(data)

    table.setStyle(

        TableStyle([

            ("BACKGROUND",(0,0),(-1,0),colors.grey),

            ("TEXTCOLOR",(0,0),(-1,0),colors.white),

            ("GRID",(0,0),(-1,-1),1,colors.black),

            ("BACKGROUND",(0,1),(-1,-1),colors.beige),

            ("ALIGN",(1,1),(-1,-1),"CENTER")

        ])

    )

    elements.append(table)

    doc.build(elements)

    return send_file(

        filename,

        as_attachment=True

    )


# ==========================================================
# EXPORT SALES PDF
# ==========================================================

@export_bp.route(
    "/export/sales-pdf",
    methods=["GET"]
)
def export_sales_pdf():

    rows = fetch_all("""

        SELECT

            id,

            total,

            payment_method,

            created_at

        FROM orders

        ORDER BY created_at DESC

    """)

    filename = os.path.join(

        EXPORT_FOLDER,

        "sales_report.pdf"

    )

    doc = SimpleDocTemplate(

        filename,

        pagesize=A4

    )

    elements = [

        Paragraph(

            "<b>Sales Report</b>",

            styles["Title"]

        )

    ]

    data = [[

        "Order",

        "Total",

        "Payment",

        "Date"

    ]]

    for row in rows:

        data.append([

            row["id"],

            row["total"],

            row["payment_method"],

            row["created_at"]

        ])

    table = Table(data)

    table.setStyle(

        TableStyle([

            ("GRID",(0,0),(-1,-1),1,colors.black),

            ("BACKGROUND",(0,0),(-1,0),colors.lightblue),

            ("ALIGN",(0,0),(-1,-1),"CENTER")

        ])

    )

    elements.append(table)

    doc.build(elements)

    return send_file(

        filename,

        as_attachment=True

    )


# ==========================================================
# EXPORT INVENTORY PDF
# ==========================================================

@export_bp.route(
    "/export/inventory-pdf",
    methods=["GET"]
)
def export_inventory_pdf():

    rows = fetch_all("""

        SELECT

            name,

            price,

            stock

        FROM products

        ORDER BY name

    """)

    filename = os.path.join(

        EXPORT_FOLDER,

        "inventory_report.pdf"

    )

    doc = SimpleDocTemplate(

        filename,

        pagesize=A4

    )

    elements = [

        Paragraph(

            "<b>Inventory Report</b>",

            styles["Title"]

        )

    ]

    data = [[

        "Product",

        "Price",

        "Stock"

    ]]

    for row in rows:

        data.append([

            row["name"],

            row["price"],

            row["stock"]

        ])

    table = Table(data)

    table.setStyle(

        TableStyle([

            ("GRID",(0,0),(-1,-1),1,colors.black),

            ("BACKGROUND",(0,0),(-1,0),colors.lightgreen),

            ("ALIGN",(0,0),(-1,-1),"CENTER")

        ])

    )

    elements.append(table)

    doc.build(elements)

    return send_file(

        filename,

        as_attachment=True

    )
# ==========================================================
# EXPORT DASHBOARD REPORT (EXCEL)
# ==========================================================

@export_bp.route(
    "/export/dashboard",
    methods=["GET"]
)
def export_dashboard():

    summary = fetch_one("""

        SELECT

            COUNT(*) AS total_orders,

            IFNULL(SUM(total),0) AS total_sales,

            IFNULL(SUM(gst),0) AS total_gst,

            IFNULL(AVG(total),0) AS average_bill

        FROM orders

        WHERE payment_status='Paid'

    """)

    wb = Workbook()

    ws = wb.active

    ws.title = "Dashboard"

    ws.append(["CafeSync Dashboard Report"])

    ws.append([])

    ws.append(["Total Orders", summary["total_orders"]])

    ws.append(["Total Sales", summary["total_sales"]])

    ws.append(["Total GST", summary["total_gst"]])

    ws.append(["Average Bill", summary["average_bill"]])

    filename = os.path.join(

        EXPORT_FOLDER,

        f"dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    )

    wb.save(filename)

    return send_file(

        filename,

        as_attachment=True
    )


# ==========================================================
# EXPORT GST REPORT
# ==========================================================

@export_bp.route(
    "/export/gst",
    methods=["GET"]
)
def export_gst():

    rows = fetch_all("""

        SELECT

            id,

            total,

            gst,

            created_at

        FROM orders

        WHERE payment_status='Paid'

        ORDER BY created_at DESC

    """)

    wb = Workbook()

    ws = wb.active

    ws.title = "GST Report"

    ws.append([

        "Order ID",

        "Bill Amount",

        "GST",

        "Date"

    ])

    total_gst = 0

    for row in rows:

        ws.append([

            row["id"],

            row["total"],

            row["gst"],

            row["created_at"]

        ])

        total_gst += row["gst"]

    ws.append([])

    ws.append([

        "",

        "Total GST",

        total_gst

    ])

    filename = os.path.join(

        EXPORT_FOLDER,

        f"gst_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    )

    wb.save(filename)

    return send_file(

        filename,

        as_attachment=True

    )


# ==========================================================
# EXPORT BEST SELLING PRODUCTS
# ==========================================================

@export_bp.route(
    "/export/best-products",
    methods=["GET"]
)
def export_best_products():

    rows = fetch_all("""

        SELECT

            p.name,

            SUM(oi.quantity) AS quantity,

            SUM(oi.quantity * oi.price) AS revenue

        FROM order_items oi

        JOIN products p

        ON oi.product_id=p.id

        GROUP BY p.id

        ORDER BY quantity DESC

    """)

    wb = Workbook()

    ws = wb.active

    ws.title = "Best Selling Products"

    ws.append([

        "Product",

        "Quantity Sold",

        "Revenue"

    ])

    for row in rows:

        ws.append([

            row["name"],

            row["quantity"],

            row["revenue"]

        ])

    filename = os.path.join(

        EXPORT_FOLDER,

        f"best_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    )

    wb.save(filename)

    return send_file(

        filename,

        as_attachment=True

    )


# ==========================================================
# EXPORT REVENUE REPORT
# ==========================================================

@export_bp.route(
    "/export/revenue",
    methods=["GET"]
)
def export_revenue():

    rows = fetch_all("""

        SELECT

            DATE(created_at) AS day,

            COUNT(*) AS orders,

            SUM(total) AS revenue

        FROM orders

        WHERE payment_status='Paid'

        GROUP BY DATE(created_at)

        ORDER BY day DESC

    """)

    wb = Workbook()

    ws = wb.active

    ws.title = "Revenue"

    ws.append([

        "Date",

        "Orders",

        "Revenue"

    ])

    for row in rows:

        ws.append([

            row["day"],

            row["orders"],

            row["revenue"]

        ])

    filename = os.path.join(

        EXPORT_FOLDER,

        f"revenue_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    )

    wb.save(filename)

    return send_file(

        filename,

        as_attachment=True

    )
# ==========================================================
# EXPORT STATUS
# ==========================================================

@export_bp.route(
    "/export/status",
    methods=["GET"]
)
def export_status():

    return jsonify({

        "success": True,

        "module": "Export",

        "status": "Running",

        "version": "1.0.0"

    })


# ==========================================================
# EXPORT HEALTH
# ==========================================================

@export_bp.route(
    "/export/health",
    methods=["GET"]
)
def export_health():

    total_products = fetch_one("""

        SELECT COUNT(*) AS total

        FROM products

    """)

    total_orders = fetch_one("""

        SELECT COUNT(*) AS total

        FROM orders

    """)

    return jsonify({

        "success": True,

        "database": "Connected",

        "products": total_products["total"],

        "orders": total_orders["total"],

        "health": "Good"

    })


# ==========================================================
# EXPORT INFORMATION
# ==========================================================

@export_bp.route(
    "/export/info",
    methods=["GET"]
)
def export_info():

    return jsonify({

        "application": "CafeSync POS",

        "module": "Export",

        "supported_formats": [

            "Excel (.xlsx)",

            "PDF"

        ],

        "available_exports": [

            "Products",

            "Categories",

            "Inventory",

            "Daily Sales",

            "Monthly Sales",

            "Orders",

            "Invoices",

            "Dashboard",

            "GST",

            "Revenue",

            "Best Selling Products"

        ]

    })


# ==========================================================
# EXPORT HISTORY
# ==========================================================

@export_bp.route(
    "/export/history",
    methods=["GET"]
)
def export_history():

    files = []

    if os.path.exists(EXPORT_FOLDER):

        for file in sorted(

            os.listdir(EXPORT_FOLDER),

            reverse=True

        ):

            path = os.path.join(

                EXPORT_FOLDER,

                file

            )

            files.append({

                "file": file,

                "size_kb": round(

                    os.path.getsize(path) / 1024,

                    2

                ),

                "created":

                    datetime.fromtimestamp(

                        os.path.getctime(path)

                    ).strftime(

                        "%Y-%m-%d %H:%M:%S"

                    )

            })

    return jsonify({

        "success": True,

        "files": files

    })


# ==========================================================
# DELETE EXPORT FILE
# ==========================================================

@export_bp.route(
    "/export/delete/<filename>",
    methods=["DELETE"]
)
def delete_export(filename):

    filepath = os.path.join(

        EXPORT_FOLDER,

        filename

    )

    if not os.path.exists(filepath):

        return jsonify({

            "success": False,

            "message": "File not found"

        }), 404

    os.remove(filepath)

    return jsonify({

        "success": True,

        "message": "Export deleted successfully"

    })


# ==========================================================
# DELETE ALL EXPORTS
# ==========================================================

@export_bp.route(
    "/export/clear",
    methods=["DELETE"]
)
def clear_exports():

    count = 0

    if os.path.exists(EXPORT_FOLDER):

        for file in os.listdir(EXPORT_FOLDER):

            os.remove(

                os.path.join(

                    EXPORT_FOLDER,

                    file

                )

            )

            count += 1

    return jsonify({

        "success": True,

        "deleted_files": count

    })


# ==========================================================
# EXPORT PING
# ==========================================================

@export_bp.route(
    "/export/ping",
    methods=["GET"]
)
def export_ping():

    return jsonify({

        "status": "OK"

    })


# ==========================================================
# END OF EXPORT MODULE
# ==========================================================